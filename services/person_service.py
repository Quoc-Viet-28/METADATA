import asyncio

import unidecode
from fastapi import HTTPException, status, File
from openpyxl.reader.excel import load_workbook
from openpyxl_image_loader import SheetImageLoader

from app.constants.person_enum import TypeImagePersonEnum
from app.dto.person_dto import PersonCreate, PersonUpdate
from app.models.person_model import Person
from app.models.type_person_model import TypePerson
from app.models.ward_model import Ward
from app.services.person_camera_service import person_camera_services
from app.utils.common_utils import get_company
from app.utils.image_handle import ImageHandle
from app.utils.minio_utils import MinioServices
from bson import ObjectId
from io import BytesIO
from PIL import Image
import xlsxwriter
from fastapi.responses import StreamingResponse
import pandas as pd

class PersonService:

    def check_face(self, image):
        imageHandle = ImageHandle()
        if image is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ảnh không hợp lệ"
            )
        face = imageHandle.save_face(image)
        if face is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Lỗi upload ảnh"
            )
        return face

    async def create(self, request: PersonCreate, user):
        imageHandle = ImageHandle()
        minio = MinioServices()
        company = await get_company(user, request.id_company)
        ward = None
        if request.id_ward:
            ward = await Ward.get(request.id_ward)
            if not ward:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Không tìm thấy xã/phường"
                )
        type_person = None
        if request.id_type_person:
            type_person = await TypePerson.find_one(
                {"_id": request.id_type_person, "company": {"$ref": "Company", "$id": company.id}})
            if not type_person:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Không tìm thấy loại person"
                )
        name_search = unidecode.unidecode(request.name).lower()
        id = ObjectId()
        name = f"{company.id}/FACE/{str(id)}.jpg"
        if request.type_image == TypeImagePersonEnum.BASE64:
            image = await asyncio.to_thread(imageHandle.decode_base64_to_image, request.image)
            t = await asyncio.to_thread(imageHandle.convert_image_to_byte, image)
            face = BytesIO(t)
        else:
            image = await asyncio.to_thread(imageHandle.decode_url_to_image, request.image)
            face = await asyncio.to_thread(self.check_face, image)

        try:
            url = await minio.upload_file_from_bytesIO(face, name)
            new_data = Person(
                **request.dict(
                    exclude={"id_company", "id_ward", "is_add_all_camera", "type_image", "image", "id_type_person"}),
                company=company,
                name_search=name_search,
                ward=ward,
                image=url,
                id=id,
                type_person=type_person,
            )
            data_save = await new_data.insert()
            if request.is_add_all_camera:
                asyncio.create_task(person_camera_services.add_one_person_to_all_camera(data_save.id))
            return data_save
        except Exception as e:
            print(e)
            await asyncio.to_thread( minio.delete_file, name)
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Lỗi không xác định"
            )

    async def update(self, request: PersonUpdate, id_person, user):
        imageHandle = ImageHandle()
        minio = MinioServices()
        person = await Person.get(id_person, fetch_links=True)
        if not person:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Không tìm thấy người"
            )
        company = await get_company(user, person.company.id)

        if request.id_ward:
            ward = await Ward.get(request.id_ward)
            if not ward:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Không tìm thấy xã/phường"
                )
            person.ward = ward

        if request.id_type_person:
            type_person = await TypePerson.find_one(
                {"_id": request.id_type_person, "company": {"$ref": "Company", "$id": company.id}})
            if not type_person:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Không tìm thấy loại person"
                )
            person.type_person = type_person

        if request.name:
            person.name_search = unidecode.unidecode(request.name).lower()

        update_data = request.dict(exclude_unset=True,
                                   exclude={"id_company", "id_ward", "is_add_all_camera", "type_image", "image",
                                            "id_type_person"})
        for field, value in update_data.items():
            setattr(person, field, value)

        if request.image:
            name = f"{person.company.id}/FACE/{person.id}.jpg"
            if request.type_image == TypeImagePersonEnum.BASE64:
                image = await asyncio.to_thread(imageHandle.decode_base64_to_image, request.image)
                t = await asyncio.to_thread(imageHandle.convert_image_to_byte, image)
                face = BytesIO(t)
            else:
                image = await asyncio.to_thread(imageHandle.decode_url_to_image, request.image)
                face = await asyncio.to_thread(self.check_face, image)

            url = await minio.upload_file_from_bytesIO(face, name)
            person.image = url

        await person.replace()
        asyncio.create_task(person_camera_services.update_person_camera(person, True if request.image else False))
        return person

    async def get_all(self, id_company=None, keyword=None, type_person=None, page=0, size=10, user=None):
        company = await get_company(user, id_company)
        if type_person == "NOT_TYPE_PERSON":
            id_type_person = None
        else:
            id_type_person = ObjectId(type_person)
        return await Person.find({"company": {"$ref": "Company", "$id": company.id}},
                                 {} if keyword is None else {"name_search": {"$regex": keyword, "$options": "i"}},
                                 {} if type_person is None else Person.type_person.id == id_type_person
                                 , fetch_links=True, nesting_depths_per_field={"type_person": 1, "company": 0
                                                                               }).skip(page * size).limit(size).sort(
            "-created_at").to_list()


    async def get_count(self, id_company=None, keyword=None, type_person=None, user=None):
        company = await get_company(user, id_company)
        if type_person == "NOT_TYPE_PERSON":
            id_type_person = None
        else:
            id_type_person = ObjectId(type_person)
        return await Person.find(Person.company.id == company.id,
                                 {} if keyword is None else {"name_search": {"$regex": keyword, "$options": "i"}},
                                 {} if type_person is None else Person.type_person.id == id_type_person
                                 , fetch_links=True, nesting_depth=1).count()

    async def process_excel_data(self, excel_data, mapping):
        data_mapping = []
        for _, row in excel_data.iterrows():
            mapped_row = {key: row[header] for key, header in mapping.items()}
            data_mapping.append(mapped_row)
        return data_mapping


    async def handle_errors(self, excel_data, image_cache, image_column, sheet):
        output = BytesIO()
        excel_data = excel_data.fillna('')
        excel_data.replace([float('inf'), -float('inf')], 'Infinity', inplace=True)

        workbook_xlsx = xlsxwriter.Workbook(output, {'in_memory': True, 'nan_inf_to_errors': True})
        worksheet_xlsx = workbook_xlsx.add_worksheet()

        for idx, img_cache in image_cache.items():
            # img_width, img_height = Image.open(img_cache).size
            img_resized = Image.open(img_cache).resize((320, 320))
            img_byte_arr = BytesIO()
            img_resized.save(img_byte_arr, format='PNG')
            img_byte_arr.seek(0)
            image_column_letter = sheet.cell(row=1, column=image_column + 1).column_letter
            cell_row_err = idx + 2
            image_key = f"{image_column_letter}{cell_row_err}"
            worksheet_xlsx.set_column(f'{image_column_letter}:{image_column_letter}', 320/7)
            worksheet_xlsx.set_row(idx + 1, 320*0.75)
            worksheet_xlsx.insert_image(image_key,'',{'image_data': img_byte_arr})
        for col_num_err, header_err in enumerate(excel_data.columns):
            worksheet_xlsx.write(0, col_num_err, header_err)

        for row_num_err, row_data_err in excel_data.iterrows():
            for col_num_err, cell_data_err in enumerate(row_data_err):
                worksheet_xlsx.write(row_num_err + 1, col_num_err, cell_data_err)

        workbook_xlsx.close()
        output.seek(0)
        return output

    async def download_template_add_user(self):
        output = BytesIO()
        workbook_xlsx = pd.ExcelWriter(output, engine='xlsxwriter')
        worksheet_xlsx = workbook_xlsx.book.add_worksheet("template")
        headers = [field for field in PersonCreate.__fields__.keys()]
        print(headers)
        for col_num, header in enumerate(headers):
            worksheet_xlsx.write(0, col_num, header)
        for col_num in range(len(headers)):
            worksheet_xlsx.set_column(0, col_num, 30)
        workbook_xlsx.close()
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=template.xlsx"}
        )
    async def create_person_instance(self, row_data: dict) -> PersonCreate:
        return PersonCreate(
            name=row_data["name"],
            id_company=row_data["id_company"],
            address=row_data.get("address"),
            id_ward=row_data.get("id_ward"),
            id_type_person=row_data.get("id_type_person"),
            sex=row_data.get("sex"),
            image=row_data.get("image"),
            type_image=row_data.get("type_image"),
            is_add_all_camera=row_data.get("is_add_all_camera"),
            other_info=row_data.get("other_info"),
            birthday=row_data.get("birthday")
        )
    async def process_excel_and_generate_mapping(self, contents: bytes, mapping_data: str = None):
        excel_data = pd.read_excel(BytesIO(contents), engine='openpyxl')
        header_data_excel = excel_data.columns.to_list()
        header_person_create = [field for field in PersonCreate.__annotations__]
        unmatched_headers = [header for header in header_data_excel if header not in header_person_create]
        workbook = load_workbook(filename=BytesIO(contents))
        sheet = workbook.active
        mapping = eval(mapping_data) if mapping_data else {}
        image_loader = SheetImageLoader(sheet)
        # full_mapping = {field: header_data_excel[i] for i, field in enumerate(header_data_excel)}
        full_mapping = {key: value for key, value in mapping.items() if key in mapping}
        unmatched_headers_mapping = [value for value in mapping.values() if value not in header_data_excel]
        if unmatched_headers_mapping:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Không tìm thấy các header: {', '.join(unmatched_headers_mapping)}"
            )
        full_mapping.update(mapping)
        data_mapping = await person_service.process_excel_data(excel_data, full_mapping)
        return excel_data, header_data_excel, header_person_create, unmatched_headers, image_loader, sheet, data_mapping, full_mapping
person_service = PersonService()
